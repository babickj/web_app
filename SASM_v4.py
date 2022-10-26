import sys
import os
import pandas as pd
import numpy as np 
# import tkinter as tk
# from tkinter import ttk
# import warnings
import csv
import time

from datetime import datetime
# from tkinter import filedialog
# from tkinter import messagebox
# from tkcalendar import Calendar
from io import BytesIO

# import rec_plugin as rp
import backend
import sasm_globals
# from tooltip import ToolTip


from openpyxl import load_workbook
# from PIL import Image, ImageTk

# Used to store location to data files
DATA_DIR = sasm_globals.DATA_DIR
DB_FILE = sasm_globals.DB_FILE
NEW_HIRE = sasm_globals.NEW_HIRE # Number of days considered a new Hire
VAR_TO_TARGET = sasm_globals.VAR_TO_TARGET # Allowed variance to target
SUSP_ABS = sasm_globals.SUSP_ABS
P_DAYS = sasm_globals.periods['1st'][1]

class Billable_Hour_Tracker:

    """
    A class representing a notebook of data tables
        
    Attribute df_bench: The data file of all employees on the bench
    Invariant: df_bench is a pandas DataFrame

    Attribute df_labor: The data file of all employees on the bench
    Invariant: df_labor is a pandas DataFrame
    """  

    # Constructor
    def __init__(self):

        self.df_bench = None
        self.df_labor = None
        self.df_last_sasm = None
        self.df_consolidated = None
        self.df_final_report = pd.DataFrame()
        self.df_watch_list = None
        self.df_new_hire = None
        self.df_last_bench = None
        self.df_last_watch = None
        self.var_target = VAR_TO_TARGET
        self.new_hire = NEW_HIRE

    # Destructor
    def __del__(self):
        pass
        # close_btn()


class User_Data:
    """
    A class representing a user data for future use.

    Attribute df_bench: The data file of all employees on the bench
    Invariant: df_bench is a pandas DataFrame

    Attribute df_labor: The data file of all employees on the bench
    Invariant: df_labor is a pandas DataFrame
    """

    # Constructor
    def __init__(self):

        self.bench_path = None
        self.labor_path = None
        self.last_sasm = None
        self.home_path = None
        self.last_bench = None
        self.new_hire = None
        self.file_path = None
        self.bench_date = datetime.today()

        # Check for file and import data
        # Check to see if the user created a file before
        past_use = os.path.exists(os.path.join(DATA_DIR,DB_FILE))

        # If a file exists, upload the data
        if(past_use):
            df_out = pd.read_feather(os.path.join(DATA_DIR,DB_FILE), use_threads=True)

            try:
                self.bench_path = df_out.loc[df_out["Catagory"] == "Bench Path", "Files"].item()

                try:
                    self.bench_date = datetime.fromtimestamp(os.path.getmtime(self.bench_path)) 

                except Exception as e:
                    self.bench_date = datetime.today()

            except ValueError:
                self.bench_path = "No File Selected"

            try:
                self.labor_path = df_out.loc[df_out["Catagory"] == "Labor Path", "Files"].item()
            except ValueError:
                self.labor_path = "No File Selected"

            try:
                self.last_sasm = df_out.loc[df_out["Catagory"] == "Last SASM", "Files"].item()
            except ValueError:
                self.last_sasm = "No File Selected"

            try:
                self.home_path = df_out.loc[df_out["Catagory"] == "Home Path", "Files"].item()
            except ValueError:
                self.home_path = "No File Selected"
            
            try:
                self.last_bench = df_out.loc[df_out["Catagory"] == "Last Bench", "Files"].item()
            except ValueError: #Exception as ex:
                self.last_bench = "No File Selected"

            try:
                self.new_hire = df_out.loc[df_out["Catagory"] == "New Hire", "Files"].item()
            except ValueError: #Exception as ex:
                self.new_hire = "No File Selected"
        

        else:
            self.bench_path = "No File Selected"
            self.labor_path = "No File Selected"
            self.last_sasm = "No File Selected"
            self.last_bench = "No File Selected"
            self.new_hire = "No File Selected"
            self.home_path = "No File Selected"

    # Destructor
    def __del__(self):
        pass
        # close_btn()


def import_data(my_billable_tracker):
    """
    import_data function imports the .csv files and stores them into 
    pandas DataFrames for further data analysis.

    :param my_billable_tracker: my_billable_tracker is an object that stores
                                two pandas dataframes:
                                    1. bench data
                                    2. labor data
                                    3. combinded data (not used in this function) 
                                    4. final_bench_report (not used in the function)
                                This data is used later for analysis
 
    :return: None
    """ 
    #############################################
    #
    # CHECK THAT USER SELECTED A FILE
    #
    ##############################################
    #### CHECK TO SEE IF A VALID FILE WAS SELECTED
    if lbl_bench_path["text"] == "No File Selected":
        tk.messagebox.showwarning("No Bench Data File Selected","No Bench Data File Selected")
        return
    elif lbl_labor_path["text"] == "No File Selected":
        tk.messagebox.showwarning("No Labor Data File Selected","No Labor Data File Selected")
        return
    elif lbl_last_sasm["text"] == "No File Selected":
        tk.messagebox.showwarning("No Last SASM File Selected", "No Last SASM File Selected")
        return
    elif lbl_last_bench["text"] == "No File Selected" or lbl_last_bench["text"] == '':
        lbl_last_bench["text"] = lbl_bench_path["text"]


    #############################################
    #
    # IMPORT BENCH DATA
    #
    ##############################################
    # Try/Except ensures that the file path is valid.  If its an invalid file
    # path a warning box is displayed
    try: # Check to see if utf-8 format
        my_billable_tracker.df_bench = pd.read_csv(lbl_bench_path["text"])

    except UnicodeDecodeError:

        # If not utf-8 format, check for utf-16 format
        try:
            text = []
        
            with open(lbl_bench_path["text"], 'r', encoding='utf16') as csvf:
            
                for line in csv.reader(csvf, delimiter='\t'):
                    text.append(line)

            my_billable_tracker.df_bench = pd.DataFrame(text, columns=text[0])
            my_billable_tracker.df_bench.drop([0], inplace=True) 
            
        except Exception:
            
            tk.messagebox.showwarning("No File Found",f"Bench data file located at: {lbl_bench_path['text']} was moved or no longer exists.")
            return


    # Ensure data sources are in the correct format
    import_col = my_billable_tracker.df_bench.columns.tolist()
    chk_col = ["Empl Name & ID", "DL+Absc Hrs Variance to ITM Target %",'Month of Hire Date']
    if not all(x in import_col for x in chk_col):
        tk.messagebox.showwarning("Incorrect Bench Data Format","Bench data needs fields: 'Empl Name & ID' and 'DL+Absc Hrs Variance to ITM Target %'")
        return

    # Convert Month of hire to datetime
    my_billable_tracker.df_bench['Month of Hire Date'] = pd.to_datetime(my_billable_tracker.df_bench['Month of Hire Date'])
    my_billable_tracker.df_bench['Month of Hire Date'] = my_billable_tracker.df_bench['Month of Hire Date'].dt.strftime('%B %Y')
 
    # Turn the raw date from the .csv file into a pandas dataframe
    my_billable_tracker.df_bench = pd.DataFrame(data=my_billable_tracker.df_bench)

    #############################################
    #
    # IMPORT LAST PERIOD BENCH DATA
    #
    ##############################################
    # Try/Except ensures that the file path is valid.  If its an invalid file
    # path a warning box is displayed
    try: # Check to see if utf-8 format
        my_billable_tracker.df_last_bench = pd.read_csv(lbl_last_bench["text"])

    except UnicodeDecodeError:

        # If not utf-8 format, check for utf-16 format
        try:
            text = []
        
            with open(lbl_last_bench["text"], 'r', encoding='utf16') as csvf:
            
                for line in csv.reader(csvf, delimiter='\t'):
                    text.append(line)

            my_billable_tracker.df_last_bench = pd.DataFrame(text, columns=text[0])
            my_billable_tracker.df_last_bench.drop([0], inplace=True) 
            
        except Exception:
            
            tk.messagebox.showwarning("No File Found",f"Last bench data file located at: {lbl_last_bench['text']} was moved or no longer exists.")
            return

    #############################################
    #
    # IMPORT LABOR DATA
    #
    ############################################## 
    # Try/Except ensures that the file path is valid.  If its an invalid file
    # path a warning box is displayed
    try: # Checks to see if utf-67 encoded

        text = []
        
        with open(lbl_labor_path["text"], 'r', encoding='utf16') as csvf:
            
            for line in csv.reader(csvf, delimiter='\t'):
                 text.append(line)

        my_billable_tracker.df_labor = pd.DataFrame(text, columns=text[0])
        my_billable_tracker.df_labor.drop([0], inplace=True) 
        my_billable_tracker.df_labor['Billability Target'].replace('','0', inplace=True)
        my_billable_tracker.df_labor['DL $ Target '].replace('','0', inplace=True)

    except:

        # Try and open the file if utf-8 encoded
        try:
            my_billable_tracker.df_labor = pd.read_csv(lbl_labor_path["text"])
            my_billable_tracker.df_labor.drop([0], inplace=True) 
            my_billable_tracker.df_labor['Billability Target'].replace('','0', inplace=True)
            my_billable_tracker.df_labor['DL $ Target '].replace('','0', inplace=True)
        except:
            
            tk.messagebox.showwarning("No File Found",f"Labor data file located at: {lbl_labor_path['text']} was moved or no longer exists.")
            return
        
    # Turn the raw .csv file into a pandas dataframe
    my_billable_tracker.df_labor = pd.DataFrame(data=my_billable_tracker.df_labor)

    # Ensure data sources are in the correct format
    import_col = my_billable_tracker.df_labor.columns.tolist()
    chk_col = ["Max. Hire Dt","Emplid","Billability Variance to Target","Billability","Suspense Amount","DL $ Target ","Total Absence Amount"]
    if not all(x in import_col for x in chk_col):
        tk.messagebox.showwarning("Incorrect Labor Data Format","Labor data needs fields: 'Max. Hire Dt','Emplid','Billability Variance to Target','Billability','Suspense Amount','DL $ Target ','Total Absence Amount'")
        return
 
    #############################################
    #
    # IMPORT LAST SASM
    #
    ##############################################
    # Import the .xlsx data from the last SASM
    #Check for skipped rows in the excel file
    try:
        book = load_workbook(lbl_last_sasm["text"])
    except IOError:
        tk.messagebox.showwarning("No File Found",f"Last SASM file located at: {lbl_last_sasm['text']} was moved or no longer exists.")
        return

    ws = book.worksheets[0]
    for cell in ws["C"]:
        if cell.value is not None:
            skip_row = cell.row - 1
            break

    try:
        my_billable_tracker.df_last_sasm = pd.read_excel(lbl_last_sasm["text"], skiprows=skip_row, engine="openpyxl")
    except IOError:
        tk.messagebox.showwarning("No File Found",f"Last SASM file located at: {lbl_last_sasm['text']} was moved or no longer exists.")
        return

    my_billable_tracker.df_last_sasm = pd.DataFrame(data=my_billable_tracker.df_last_sasm)

    # Ensure data sources are in the correct format
    import_col = my_billable_tracker.df_last_sasm.columns.tolist()
    chk_col = ["SASM Notes", "SASM RM Status", "Anticipated Start Date", "Activity Comments"] 
    if not all(x in import_col for x in chk_col):
        tk.messagebox.showwarning("Incorrect Last SASM Data Format","Last SASM data needs fields: 'SASM Notes', 'SASM RM Status', and 'Anticipated Start Date'")
        return

    #############################################
    #
    # IMPORT CLEARANCE DATA
    #
    ##############################################

    # Ensure valid file path
    """
    try:
        my_billable_tracker.df_last_cleared = pd.read_csv(lbl_last_cleared["text"])
    except IOError:
        tk.messagebox.showwarning("No File Found", f"Clearance data file located at: {lbl_last_cleared['text']} was moved or no longer exists.")
        return

    # Ensure data sources are in the correct format
    import_col = my_billable_tracker.df_last_cleared.columns.tolist()
    chk_col = ['Emplid', 'Clearance']
    if not all(x in import_col for x in chk_col):
        tk.messagebox.showwarning("Incorrect Clearance Data Format","Clearance data needs fields: 'Emplid' and 'Clearance'")
        return

    # Turn raw data from .csv file inat a dataframe
    my_billable_tracker.df_last_cleared = pd.DataFrame(data=my_billable_tracker.df_last_cleared)
    """

    #############################################
    #
    # IMPORT NEW HIRE DATA
    #
    ##############################################
    # Import the .xlsx data from the hiring roster
    #Check for skipped rows in the excel file
    # First get new hire data from the new hire .csv
    try:
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            book = load_workbook(lbl_new_hire["text"])
    except IOError:
        tk.messagebox.showwarning("No File Found",f"Employee Roster file located at: {lbl_new_hire['text']} was moved or no longer exists.")
        return

    ws = book.worksheets[0]
    for cell in ws["C"]:
        if cell.value is not None:
            skip_row = cell.row - 1
            break
    # Ensure valid file path
    try:
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            my_billable_tracker.df_new_hire = pd.read_excel(lbl_new_hire["text"], skiprows=skip_row, engine="openpyxl")
    except IOError:
        tk.messagebox.showwarning("No File Found", f"New Hire data file located at: {lbl_new_hire['text']} was moved or no longer exists.")

    # Ensure data sources are in the correct format
    import_col = my_billable_tracker.df_new_hire.columns.tolist()
    chk_col = ['Employee ID', 'Hire Date']
    
    if not all(x in import_col for x in chk_col):
        tk.messagebox.showwarning("Incorrect Hire Date Format","Hire data needs fields: 'Employedd ID' and 'Hire Date'")
        return
    my_billable_tracker.df_new_hire = pd.DataFrame(data=my_billable_tracker.df_new_hire)
    my_billable_tracker.df_new_hire = my_billable_tracker.df_new_hire.set_index("Employee ID")
    
    # Now get comments from the last weeks SASM meeting
    try:
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            book = load_workbook(lbl_last_sasm["text"])

            ws = book["New Hires"]
            for cell in ws["C"]:
                if cell.value is not None:
                    skip_row = cell.row - 1
                    break

            tmp_df = pd.read_excel(lbl_last_sasm["text"], sheet_name="New Hires", 
                                    skiprows=skip_row, engine="openpyxl")

            tmp_df = tmp_df.set_index("Employee ID")
            tmp_df.index = tmp_df.index.fillna(0)
            tmp_df.index = tmp_df.index.astype(int)
            
            for idx1 in my_billable_tracker.df_new_hire.index:
                for idx2 in tmp_df.index:
                    if idx1 == idx2:
                        
                        tmp_poc = tmp_df.loc[tmp_df.index == idx2]["SASM POC"]
                        my_billable_tracker.df_new_hire.loc[my_billable_tracker.df_new_hire.index == idx1, "SASM POC"] = tmp_poc 
            

    except Exception as e:
        pass


    
    ##############################################
    #
    # IMPORT LAST WEEK'S WATCH LIST
    #
    ##############################################

    try:
        book = load_workbook(lbl_last_sasm["text"])
    except IOError:
        tk.messagebox.showwarning("No File Found",f"Last SASM file located at: {lbl_last_sasm['text']} was moved or no longer exists.")
        return

    try:
        ws = book["Watch List"]
        for cell in ws["C"]:
            if cell.value is not None:
                skip_row = cell.row - 1
                break
        my_billable_tracker.df_last_watch = pd.read_excel(lbl_last_sasm["text"], sheet_name="Watch List"
                                            ,skiprows=skip_row, engine="openpyxl")
    except Exception as e:
        my_billable_tracker.df_last_watch = None
  
    ##########################################
    #
    # DISPLAY IMPORTED DATA IN TABS
    # 
    ##########################################
    
    # UPDATE BENCH DATA
    tab1_txt.config(state='normal')
    tab1_txt.delete('1.0', 'end')
    tab1_txt.insert('1.0',f"\nBench List (Raw Data): \n\n {my_billable_tracker.df_bench}") 
    tab1_txt.config(state='disabled')
    # UPDATE LABOR
    tab2_txt.config(state='normal')
    tab2_txt.delete('1.0', 'end')
    tab2_txt.insert('1.0',f"\nLabor List (Raw Data): \n\n {my_billable_tracker.df_labor}") 
    tab2_txt.config(state='disabled')


def generate_report(my_billable_tracker, user):

    """
    generate_report function does ALL the data analysis.  It uses the bench and labor 
    dataframes and runs the following deciion tree:
        1. If member is on bench list verify:
            a. No approved leave
            b. No suspended charges
            c. Not in a non-billable billet
            d. Not a hew hire (< 6 months)
        2. If member is not on bench list, check associated labor data and verify:
            a. Member billed required hours, s.t. constraints in paragraph 1.

        The results are stored in the final_bench_report dataframe contained in
        the my_billable_tracker object

    :param my_billable_tracker: my_billable_tracker is an object that stores
                                two pandas dataframes:
                                    1. bench data
                                    2. labor data
                                    3. final_bench_report
 
    :return: None
    """

    # Ensure the data was inputed prior to generating the report
    import_data(my_billable_tracker)

    # Pre-process the data, e.g., index, types, etc
    backend.pre_process(my_billable_tracker)
    
    #########################################
    #
    # RUN DECISION TREE
    #
    #########################################

    # 1st decition point
    # Is bench data < 4 business days of this reporting period
    use_bench_only = backend.dt_less_4_bd(my_billable_tracker,user)
    # use_bench_only = False
    
    # If we pass dt point 1, continue with the rest of the tree
    if use_bench_only:
        tk.messagebox.showinfo(title="Using Bench Data Only", message=f"Within {P_DAYS} days of reporting period, only using data from {user.bench_path}.")
        # print("Use Bench Only")
        pass

    else:

        # This decision point check to see if employee is billabile
        backend.dt_isBillable(my_billable_tracker)

        # This decision points checks to see(SUSP + ABS > 0.75) and (NOT on prior period bench)
        backend.dt_pt_2(my_billable_tracker)

    # (DL + ABS HRS VAR to ITM target % < -0.50) && (Billability Target > 0.50)?
    backend.dt_pt_3(my_billable_tracker)
    
    # If employee was on official SASM, do NOT move to watchlist
    backend.dt_pt_4(my_billable_tracker)


    #########################################
    #
    # UPDATE GUI
    #
    #########################################
    # Remove Senior Associates and Principals
    my_billable_tracker.df_consolidated.loc[(my_billable_tracker.df_consolidated['Level '] == "Senior Associate"), "Notes"] = "Senior Associate - Removed from tracker"
    my_billable_tracker.df_consolidated.loc[(my_billable_tracker.df_consolidated['Level '] == "Principal/Director"), "Notes"] = "Princial - Removed from tracker"
    my_billable_tracker.df_consolidated.loc[(my_billable_tracker.df_consolidated['Level '] == "Senior Associate"), "on_bench"] = False
    my_billable_tracker.df_consolidated.loc[(my_billable_tracker.df_consolidated['Level '] == "Principal/Director"), "on_bench"] = False

    # Generate the final bench report and watchlist
    my_billable_tracker.df_final_report = my_billable_tracker.df_consolidated.loc[my_billable_tracker.df_consolidated['on_bench'] == True]
    my_billable_tracker.df_watch_list = my_billable_tracker.df_consolidated.loc[my_billable_tracker.df_consolidated['on_bench'] == "watch"]

    # Scrub 'Senior Associate' and "Principal/Director"
    my_billable_tracker.df_consolidated = my_billable_tracker.df_consolidated.loc[my_billable_tracker.df_consolidated['Level '] != "Senior Associate"]
    my_billable_tracker.df_consolidated = my_billable_tracker.df_consolidated.loc[my_billable_tracker.df_consolidated['Level '] != "Principal/Director"]

    # Catagorize the new hires
    my_billable_tracker.df_new_hire = my_billable_tracker.df_new_hire.loc[my_billable_tracker.df_new_hire["Hire Date"] + pd.DateOffset(days=NEW_HIRE) >= pd.Timestamp.today()]

    # Update column names to match desiered output.
    # Rename columns to match previous report
    my_billable_tracker.df_new_hire = my_billable_tracker.df_new_hire.rename(columns={"Full Legal Name":"Name", "Job Leader":"CM as of Hire Date"})

    cols = ["Name", "Hire Date", "Job Requisition","CM as of Hire Date", "Principal", "SASM POC", 
            "Management Level", "Location", "Opportunities (Contract field)", "Requisition Type", "Job Family",
             "Clearance as confirmed by CM"
    ]
    my_billable_tracker.df_new_hire = my_billable_tracker.df_new_hire.reindex(columns=cols)

    # Check to see if there is new hire data in the previous SASM. 
    # If there is, import the old data
    try:
        book = load_workbook(lbl_last_sasm["text"])
        ws = book["New Hires"]

        # Check to see if there are blank rows to skip
        for cell in ws["C"]:
            if cell.value is not None:
                skip_row = cell.row - 1
                break
        
        # Load ws into temp df
        tmp_df = pd.read_excel(lbl_last_sasm["text"], sheet_name="New Hires", 
                                skiprows=skip_row, engine="openpyxl")
        tmp_df = pd.DataFrame(data=tmp_df)
        tmp_df.set_index("Employee ID", inplace=True)

        # Copy applicable data from last_sasm to new hires df
        my_billable_tracker.df_new_hire.loc[my_billable_tracker.df_new_hire.index.isin(tmp_df.index), 
                                            "Job Requisition"] = tmp_df['Job Requisition']
        my_billable_tracker.df_new_hire.loc[my_billable_tracker.df_new_hire.index.isin(tmp_df.index), 
                                            "Opportunities (Contract field)"] = tmp_df["Opportunities (Contract field)"]
        my_billable_tracker.df_new_hire.loc[my_billable_tracker.df_new_hire.index.isin(tmp_df.index),                      
                                            "Requisition Type"] = tmp_df["Requisition Type"]
        my_billable_tracker.df_new_hire.loc[my_billable_tracker.df_new_hire.index.isin(tmp_df.index), 
                                            "Clearance as confirmed by CM"] = tmp_df["Clearance as confirmed by CM"]
    
        # my_billable_tracker.df_new_hire.to_csv("dump.csv")
    except Exception as e:
        # print(e)
        tk.messagebox.showwarning("No 'New Hires' tab in previous SASM reprt", "No 'New Hires' tab in previous SASM reprt")
     

    #########################################
    #
    # DISPLAY GENERATED REPORT TO USER
    #
    #########################################
    tab3_txt.config(state='normal')
    tab3_txt.delete(1.0, 'end')
    tab3_txt.insert('1.0', f"\nGenerated Bench Report:\n\n{my_billable_tracker.df_final_report}")
    tab3_txt.config(state='disabled')

    # Bring report tab into focus
    tabControl.select(tab3)
    
    # Create file / used for web-app only
    # export_btn(mbt,user)


def pair_openings():

    # Call pair openings from recommender
    match = rp.main()

    #for key, val in match.items():
    #    print(f"{key}: {val}")

    tab4_txt.config(state='normal')
    tab4_txt.delete('1.0', 'end')
    tab4_txt.insert('1.0', "\nEMPLOYEE MATCHES:\n")
    for key, val in match.items():
        tab4_txt.insert('end', '\n')
        tab4_txt.insert('end', f"{key}: {val}")
    
    tab4_txt.config(state='disabled')

    # Bring employee tab into focus
    tabControl.select(tab4)


def close_btn():

    """
    clost_btn function does two things to ensure a controlled closing:
        1. Saves all user data
        2. Closes the program
        
    :params: None
            
    :return: None
    """    
  
    # Save the user data into a pickle file
    out_path = DB_FILE

    df_out = None 

    # Create dataframe to hold the user file
    df_out = {"Catagory": ["Bench Path", "Labor Path", "Last SASM", "Last Bench", "New Hire", "Home Path"],
                "Files":[ lbl_bench_path["text"], lbl_labor_path["text"], lbl_last_sasm["text"], lbl_last_bench["text"], lbl_new_hire["text"],os.path.curdir]
            }


    df_out = pd.DataFrame(data=df_out)

    # Save to hard-drive
    df_out.to_feather(out_path)

    sys.exit()


def export_btn(mbt,user):

    """
    export_btn function allows the user to export the 
    bench report to a formatted Excel file.  The format of
    the Excel is hard-coded.  Need to add user format in future
    release.
        
    :params: my_billable_tracker object (mbt)
            
    :return: None
    """  
    ############################################
    #
    #  GENERATES THE BENCH REPORT MAIN SHEET
    #
    ############################################

    # Hold the export file name
    f_name = "Bench_Report.xlsx"
    sheet_name = "Bench_Report"

    # Prevents any data output from being interpretted as formulas
    options = {}
    options['strings_to_formulas'] = False
    options['strings_to_urls'] = False
    writer = pd.ExcelWriter(os.path.join(DATA_DIR, f_name), engine='xlsxwriter', engine_kwargs={'options':options})
    
    # Format the dataframe to the desired format
    cols = ["Empl Name & ID", "Level ", "Billability Target", "Billability Variance to Target", "Inc or Curr SA 123", "Inc or Curr Principal 123",
               "Curr VP ", "Bench Mos", "SASM RM Status", "SASM Notes", "Anticipated Start Date", "RM Status", "Activity Comments", "Clearance Eligibility ",
               "IMT", "Pool", "Job Family ", "Proximity Location ", "Month of Hire Date", "Admin (BM) Hrs", "Tech Mktg Hrs", "BMMA Hrs", "Sect Admin Hrs", "Mktg Hrs",
               "Sect Dev Hrs", "Admin (SS) Hrs", "B&P Hrs", "IR&D Hrs", "Firmwide Inv Hrs", "Special Proj Hrs", "Total Unallowable Hrs", "Suspense Hrs", "Billed Hrs + Abs",
               "DL+Absc Hrs ITM Target", "DL+Absc Hrs Variance to ITM Target", "DL+Absc Hrs Variance to ITM Target %", "YTD Bench Impact", "+1 Month FT2 DL Hrs",
               "+2 Month FT2 DL Hrs", "+3 Month FT2 DL Hrs"]

    # Verify that the final report was generated prior to exporting
#    if mbt.df_final_report is not pd.DataFrame:
#        generate_report(mbt,user)

    # Populate the dataframe for the final report
    mbt.df_final_report = mbt.df_final_report.reindex(columns=cols)

    # Rename the columns to match the desired output
    mbt.df_final_report = mbt.df_final_report.rename(columns={"Level ":"Curr Level", "Billability Variance to Target":"Variance", "Inc or Curr SA 123":"Emp Sr Associate",
                                         "Inc or Curr Principal 123":"Curr Principal"})

    # Update "Month of Hire Date" to pd.Datetime
    mbt.df_final_report ['Month of Hire Date']= pd.to_datetime(mbt.df_final_report['Month of Hire Date'])


    cols = mbt.df_final_report.columns.values.tolist()

    # Outputs bench report to excel
    mbt.df_final_report.to_excel(writer, sheet_name, index=True, startrow = 2)

    # Add title to the Excel file
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    worksheet.write(0,0,"Bench Report as of "+datetime.now().strftime('%d %b %Y'), workbook.add_format({'bold':True, 'color':'blue','size':14}))

    # Add color to the table header
    header_format = workbook.add_format({'bold':True, 'text_wrap':True, 'fg_color':'#FDE9D9', 'border':1, 'align':'left', 'valign':'top'})
    cols.insert(0, "Emp ID")
    
    # Add index labels in header
    for col_num, value in enumerate(cols):
       worksheet.write(2, col_num, value, header_format)

    # Set the column width
    row_idx, col_idx = mbt.df_final_report.shape
    worksheet.set_column(0, col_idx, 20)

    # Set default Excel window size
    workbook.window_width = 30000
    workbook.window_height = 30000

    # Add the Excel table structure.
    # Create a list of column headers, to use in add_table().
    column_settings = [{'header': column} for column in cols]
    # Create the table
    worksheet.add_table(2, 0, row_idx+2, col_idx, {'columns': column_settings})

    # Add a remark the the end of the file
    worksheet.write(len(mbt.df_final_report)+4, 0, 'Remark:', workbook.add_format({'bold':True}))
    worksheet.write(len(mbt.df_final_report)+5, 0, "The last update time is: "+datetime.now().strftime('%H:%M')+'.')

    # Add conditional formating to new hire column 'T'
    format1 = workbook.add_format({'bg_color':   '#FFEB9C',
                               'font_color': '#9C6500',
                               'num_format': 'mmm-yy'})

    # Formats the new hire date column in the desired format
    format2 = workbook.add_format({'num_format': 'mmm-yy'})
    format3 = workbook.add_format({'num_format': 'mm/dd/yy'})
    format4 = workbook.add_format({'bg_color': '#04591B'})
    format5 = workbook.add_format({'bg_color': '#FFC7CE',
                               'font_color': '#9C0006',
                               'num_format': 'mm/dd/yy'})

    # Logic check to test if person is a new hire
    today = pd.Timestamp.today() - pd.DateOffset(days=NEW_HIRE)

    # Adds conditional formating rules to .xlxs file
    worksheet.conditional_format('T4:T500', {'type':     'date',
                                       'criteria': 'greater than or equal to',
                                       'value':    today,
                                       'format':   format1})

    worksheet.conditional_format('T4:T500', {'type':     'date',
                                       'criteria': 'less than',
                                       'value':    today,
                                       'format':   format2})

    worksheet.conditional_format('L4:L500', {'type':     'date',
                                       'criteria': 'greater than',
                                       'value':    pd.Timestamp.today(),
                                       'format':   format3})

    worksheet.conditional_format('L4:L500', {'type':     'date',
                                       'criteria': 'between',
                                       "minimum":   datetime(1900,1,2),
                                       'maximum':   pd.Timestamp.today(),
                                       'format':   format5})

    worksheet.conditional_format('J3:L3', {'type':  'cell',
                                        'criteria': 'not equal to',
                                        'value':    0,
                                        'format':   format4})
    
    # Ensure column J must come from a drop down list
    worksheet.data_validation('J4:J500', {'validate': 'list',
                                  'source': ['Billable', 'Not Available', 'Not Billable by Design', 'Future Available', 'Available', 'Hard-Booked']})

    # Label for new hire
    worksheet.write('T2', "New Hire", format1)
    worksheet.write('L2', "Data Passed - Please Update", format5)

    ############################################
    #
    #  GENERATES THE WATCHLIST SHEET
    #
    ############################################


    cols = ["Empl Full Name ", "Notes", "Level ", "Billability Target", "Billability Variance to Target", "Inc or Curr SA 123", "Inc or Curr Principal 123",
            "Curr VP ", "Bench Mos", "SASM RM Status", "SASM Notes", "Anticipated Start Date", "RM Status", "Activity Comments", "Clearance Eligibility ",
            "IMT", "Pool", "Job Family ", "Proximity Location ", "Month of Hire Date", "Admin (BM) Hrs", "Tech Mktg Hrs", "BMMA Hrs", "Sect Admin Hrs", "Mktg Hrs",
            "Sect Dev Hrs", "Admin (SS) Hrs", "B&P Hrs", "IR&D Hrs", "Firmwide Inv Hrs", "Special Proj Hrs", "Total Unallowable Hrs", "Suspense Hrs", "Billed Hrs + Abs",
            "DL+Absc Hrs ITM Target", "DL+Absc Hrs Variance to ITM Target", "DL+Absc Hrs Variance to ITM Target %", "YTD Bench Impact", "+1 Month FT2 DL Hrs",
            "+2 Month FT2 DL Hrs", "+3 Month FT2 DL Hrs"]

    mbt.df_watch_list = mbt.df_watch_list.reindex(columns=cols)

    # Rename the columns to match the desired output
    mbt.df_watch_list = mbt.df_watch_list.rename(columns={"Level ":"Curr Level", "Billability Variance to Target":"Variance", "Inc or Curr SA 123":"Emp Sr Associate",
                                         "Inc or Curr Principal 123":"Curr Principal"})

    # Update "Month of Hire Date" to pd.Datetime
    mbt.df_watch_list ['Month of Hire Date']= pd.to_datetime(mbt.df_watch_list['Month of Hire Date'])


    cols = mbt.df_watch_list.columns.values.tolist()

    # Outputs bench report to excel
    mbt.df_watch_list.to_excel(writer, "Watch List", index=True, startrow = 2)

    # Add title to the Excel file
    workbook = writer.book
    worksheet = writer.sheets["Watch List"]
    #worksheet.write(0,0,"Bench Report as of "+datetime.now().strftime('%d %b %Y'), workbook.add_format({'bold':True, 'color':'blue','size':14}))

    # Add color to the table header
    header_format = workbook.add_format({'bold':True, 'text_wrap':True, 'fg_color':'#FDE9D9', 'border':1, 'align':'left', 'valign':'top'})
    cols.insert(0, "Empl ID")
    
    # Add index labels in header
    for col_num, value in enumerate(cols):
       worksheet.write(2, col_num, value, header_format)

    # Set the column width
    row_idx, col_idx = mbt.df_watch_list.shape
    worksheet.set_column(0, col_idx, 20)

    # Set default Excel window size
    workbook.window_width = 30000
    workbook.window_height = 30000

    # Add the Excel table structure.
    # Create a list of column headers, to use in add_table().
    column_settings = [{'header': column} for column in cols]
    # Create the table
    worksheet.add_table(2, 0, row_idx+2, col_idx, {'columns': column_settings})

    ############################################
    #
    #  GENERATES THE NEW HIRE SHEET
    #
    ############################################

    # Output 'New Hire' to another sheet
    cols = ["Name", "Hire Date", "Job Requisition","CM as of Hire Date", "Principal", "SASM POC", 
            "Management Level", "Location", "Opportunities (Contract field)", "Requisition Type", "Job Family",
             "Clearance as confirmed by CM"
    ]

   # mbt.df_new_hire = mbt.df_new_hire.reindex(columns=cols)
    cols.insert(0, "Employee ID")
    
    # Format the hire date
#    mbt.df_new_hire["Hire Date"] = mbt.df_new_hire["Hire Date"].dt.strftime('%m/%d/%Y')

    mbt.df_new_hire.to_excel(writer, "New Hires", index=True, startrow = 2)
    worksheet = writer.sheets["New Hires"]
    column_settings = [{'header': column} for column in cols]
    # Create the table
    row_idx, col_idx = mbt.df_new_hire.shape
    worksheet.set_column(0, col_idx, 20)
    worksheet.add_table(2, 0, row_idx+2, col_idx, {'columns': column_settings})

    # Add the formatted header
    # Light red fill with dark red text.
    # Light red fill with dark red text.
    red_format = workbook.add_format({'bg_color':   '#FFC7CE',
                               'font_color': '#9C0006'})

    # Light yellow fill with dark yellow text.
    yellow_format = workbook.add_format({'bg_color':   '#FFEB9C',
                               'font_color': '#9C6500'})

    # Green fill with dark green text.
    green_format = workbook.add_format({'bg_color':   '#C6EFCE',
                               'font_color': '#006100'})
    worksheet.write('J2', "Red - Not Billable", red_format)
    worksheet.write('K2', "Yellow - Hardbooked", yellow_format)
    worksheet.write('L2', "Green - Billable or SI", green_format)

    #############################################
    #
    #  GENERATES THE RAW DATA SHEET
    #
    ############################################

    # # Output raw data to another sheet
    # cols = ["Empl Full Name ", "Notes", "Level ", "Billability Target", "Billability Variance to Target", "Inc or Curr SA 123", "Inc or Curr Principal 123",
    #            "Curr VP ", "Bench Mos", "SASM RM Status", "SASM Notes", "Anticipated Start Date", "RM Status", "Activity Comments", "Clearance Eligibility ",
    #            "IMT", "Pool", "Job Family ", "Proximity Location ", "Month of Hire Date", "Admin (BM) Hrs", "Tech Mktg Hrs", "BMMA Hrs", "Sect Admin Hrs", "Mktg Hrs",
    #            "Sect Dev Hrs", "Admin (SS) Hrs", "B&P Hrs", "IR&D Hrs", "Firmwide Inv Hrs", "Special Proj Hrs", "Total Unallowable Hrs", "Suspense Hrs", "Billed Hrs + Abs",
    #            "DL+Absc Hrs ITM Target", "DL+Absc Hrs Variance to ITM Target", "DL+Absc Hrs Variance to ITM Target %", "YTD Bench Impact", "+1 Month FT2 DL Hrs",
    #            "+2 Month FT2 DL Hrs", "+3 Month FT2 DL Hrs"
    # ]
    
    # mbt.df_consolidated = mbt.df_consolidated.reindex(columns=cols)
    # cols.insert(0, "Empl ID")
    # mbt.df_consolidated.to_excel(writer, "Raw Data", index=True, startrow = 2)
    # worksheet = writer.sheets["Raw Data"]
    # column_settings = [{'header': column} for column in cols]
    # # Create the table
    # row_idx, col_idx = mbt.df_consolidated.shape
    # worksheet.set_column(0, col_idx, 20)
    # worksheet.add_table(2, 0, row_idx+2, col_idx, {'columns': column_settings})

    writer.close()
    
    # Tell use file was exported:
    user.file_path = os.path.join(DATA_DIR, f_name)
    
#    open_now = tk.messagebox.askquestion(title="Bench Report Exported", message=f"Bench Report Exported to: {file_path}.\nWould you like to open now?")
#    if open_now == 'yes':
    # os.system("open " + file_path)
    

def select_bench_file(user, date_label):

    """
    select_bench_file function allows the user to select the 
    location of the bench .csv file.
        
    :params: None
            
    :return: None
    """   
    filename = filedialog.askopenfilename(initialdir = ".",
                                          title = "Select a File",
                                          filetypes = (("CSV files",
                                                        "*.csv"),
                                                       ("all files",
                                                        "*.*")))
      
    # Change label contents
    bench_data = filename
    lbl_bench_path["text"] = bench_data

    # Update date
    try:
        
        user.bench_date = datetime.fromtimestamp(os.path.getmtime(filename))
        tmp_date = f"{user.bench_date.month}/{user.bench_date.day}/{user.bench_date.year}"
        date_label['text'] = tmp_date

    except Exception as e:
        
        user.bench_date = datetime.today()
        tmp_date = f"{user.bench_date.month}/{user.bench_date.day}/{user.bench_date.year}"
        date_label['text'] = tmp_date

def select_sasm_file():
    """
    select_sasm_file function allows the user to select the 
    location of the last SASM .xlsx file.
        
    :params: None
            
    :return: None
    """   
    filename = filedialog.askopenfilename(initialdir = ".",
                                          title = "Select a File",
                                          filetypes = (("EXCEL files",
                                                        "*.xlsx"),
                                                       ("all files",
                                                        "*.*")))
      
    # Change label contents
    last_sasm = filename
    lbl_last_sasm["text"] = last_sasm


def select_labor_file():


    """
    select_labor_file function allows the user to select the 
    location of the labor .csv file.
        
    :params: None
            
    :return: None
    """   
    filename = filedialog.askopenfilename(initialdir = ".",
                                          title = "Select a File",
                                          filetypes = (("CSV files",
                                                        "*.csv"),
                                                       ("all files",
                                                        "*.*")))
      
    # Change label contents
    labor_data = filename
    lbl_labor_path["text"] = labor_data


def select_last_bench():
    """
    select_cleared_file function allows the user to select the 
    location of the cleared .csv file.
        
    :params: None
            
    :return: None
    """   
    lbl_last_bench["text"] = filedialog.askopenfilename(initialdir = ".",
                                          title = "Select a File",
                                          filetypes = (("CSV files",
                                                        "*.csv"),
                                                       ("all files",
                                                        "*.*")))
    
    
def select_new_hire():
    """
    select_new_hire function allows the user to select the 
    location of the new hire .csv file.
        
    :params: None
            
    :return: None
    """   
    lbl_new_hire["text"] = filedialog.askopenfilename(initialdir = ".",
                                          title = "Select a File",
                                          filetypes = (("XLSX files",
                                                        "*.xlsx"),
                                                       ("all files",
                                                        "*.*")))

### ALLOWS CODE TO BE EXECUTED BY THE OS. FACILITATES TRANSITION TO AN .EXE FILE ###
if __name__ == "__main__":

    """
    the main function is the entry point for the application.  This is where the GUI
    is set up. 
        
    :params: None
            
    :return: None
    """  

    # Import data for processing
    my_prog = User_Data()
    my_billable_tracker = Billable_Hour_Tracker()

    # Shows all Pandas DataFrame columns on the GUI
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    pd.set_option("display.expand_frame_repr", False)

    # Generate and Label a new window
    window = tk.Tk()
    window.title("SIG Anayltics Staff Management")
    
    # Size the window and layout the grid
    window.rowconfigure(0, minsize=100, weight=1)
    window.rowconfigure(1, minsize=400, weight=1)
    window.columnconfigure(1, minsize=800, weight=1)

    # DISPLAY BOOZ ALLEN LOGO
    # img_path = "BAH.jpeg"
    # image = Image.open(img_path)
    # photo = ImageTk.PhotoImage(image)
    # lbl_photo = tk.Label(window, image=photo)
    # lbl_photo.grid(row=0, column=2)

    ###################################
    #
    # CREATE BUTTONS ON LEFT HAND SIDE
    #
    ###################################

    # Frame Used to hold Buttons
    frm_buttons = tk.Frame(window, relief=tk.RAISED, bd=2)
    frm_buttons.grid(row=0, column=0, sticky="nsew")

    # Create import data button
    btn_input_data = tk.Button(frm_buttons, text="Import Data", font='Helvetica 14 bold', command=lambda : import_data(my_billable_tracker))
    btn_input_data.grid(row=1, column=0, sticky="ew", padx=50, pady=3)

    # Create the bench report generation button
    btn_generate_report = tk.Button(frm_buttons, text="Generate Bench Report", font='Helvetica 14 bold', command=lambda : generate_report(my_billable_tracker,my_prog))
    btn_generate_report.grid(row=2, column=0, sticky="ew", padx=50, pady=3)

    # Pair employee button
    # btn_pair_talent = tk.Button(frm_buttons, text="Pair Openings", font='Helvetica 14 bold', command=pair_openings)
    # btn_pair_talent.grid(row=3, column=0, sticky = "ew", padx = 50, pady=3)

    # Export bench report button
    btn_export_report = tk.Button(frm_buttons, text="Export Report", font='Helvetica 14 bold', command=lambda : export_btn(my_billable_tracker,my_prog))
    btn_export_report.grid(row=3,column=0, sticky="ew", padx=50, pady=3)

    # Close program button
    btn_close = tk.Button(frm_buttons, text="Close", font='Helvetica 14 bold', command=close_btn)
    btn_close.grid(row=4, column=0, sticky="ew", padx=50, pady=3)

    ###################################
    #
    # CREATE AREA TO SELECT DATA FILES
    # TO THE RIGHT OF BUTTONS
    #
    ###################################

    # Create the frame to show the file path to the data sources
    frame_data = tk.Frame(window, relief=tk.RAISED, bd=2)
    frame_data.grid(row=0,column=1, sticky="nsew")

    # This labels the window frame
    data_label = tk.Label(frame_data, text="Data Sources", 
                            font='Helvetica 18 bold', fg='green')
    data_label.grid(row=0, column=1, sticky="ew")
    
    # Label the bench files path
    bench_label = tk.Label(frame_data, text="Bench Data (.csv):", font='Helvetica 14 bold')
    bench_label.grid(row=1, column=0, sticky="ew")
    lbl_bench_path = tk.Label(frame_data, text=my_prog.bench_path, relief=tk.SUNKEN)
    lbl_bench_path.grid(row=1, column=1, sticky="ew")

    # Insert button to allow bench file location to be selected
    btn_bench_select = tk.Button(frame_data, text="Select File", command=lambda:select_bench_file(my_prog, date_label))
    btn_bench_select.grid(row=1, column=2, sticky="ew", padx=5, pady=5)

    # Insert label to display date of file
    tmp_date = f"{my_prog.bench_date.month}/{my_prog.bench_date.day}/{my_prog.bench_date.year}"

    date_label = tk.Label(frame_data, text=tmp_date, relief=tk.SUNKEN)
    date_label.grid(row=1, column=3, sticky="ew")

    # Add button to select new update
    btn_sel_date = tk.Button(frame_data, text=f"Change Date of Bench Data", command=lambda:backend.select_date(window, my_prog,date_label))
    btn_sel_date.grid(row=0, column=3, sticky="ew")
    
    # Place '?' icon for hover-over
    icn_label = tk.Label(frame_data, text="?", font='Helvetica 24 bold', relief=tk.RIDGE, bg='light blue')
    icn_label.grid(row=1, column=4, sticky='w')
    
    # bind the button to display hover tip
    ToolTip(widget = icn_label, text = "This updates the date \nused by the program for the bench data.")

    # Label the labor data file path
    labor_label = tk.Label(frame_data, text="Labor Data(.csv):", font='Helvetica 14 bold')
    labor_label.grid(row=2, column=0, sticky="ew")
    lbl_labor_path = tk.Label(frame_data, text = my_prog.labor_path, relief=tk.SUNKEN)
    lbl_labor_path.grid(row=2, column=1, sticky="ew")

    # Insert button to select labor file location
    btn_labor_select = tk.Button(frame_data, text="Select File", command=select_labor_file)
    btn_labor_select.grid(row=2, column=2, sticky="ew", padx=5, pady=5)

    # Label the last SASM data file path (in Excel)
    sasm_label = tk.Label(frame_data, text="Last SASM(.xlsx):", font='Helvetica 14 bold')
    sasm_label.grid(row=3, column=0, sticky="ew")
    lbl_last_sasm = tk.Label(frame_data, text = my_prog.last_sasm, relief=tk.SUNKEN)
    lbl_last_sasm.grid(row=3, column=1, sticky="ew")

    # Insert button to select last sasm file location
    btn_last_sasm = tk.Button(frame_data, text="Select File", command=select_sasm_file)
    btn_last_sasm.grid(row=3, column=2, sticky="ew", padx=5, pady=5)

    # Label the last Clearance data file path (in Excel)
    cleared_label = tk.Label(frame_data, text="Last Bench(.csv):", font='Helvetica 14 bold')
    cleared_label.grid(row=4, column=0, sticky="ew")
    lbl_last_bench = tk.Label(frame_data, text = my_prog.last_bench, relief=tk.SUNKEN)
    lbl_last_bench.grid(row=4, column=1, sticky="ew")

    # Insert button to select last cleared data file location
    btn_last_bench = tk.Button(frame_data, text="Select File", command=select_last_bench)
    btn_last_bench.grid(row=4, column=2, sticky="ew", padx=5, pady=5)

    # Label the New Hire data file path (in Excel)
    new_hire_label = tk.Label(frame_data, text="New Hire(.csv):", font='Helvetica 14 bold')
    new_hire_label.grid(row=5, column=0, sticky="ew")
    lbl_new_hire = tk.Label(frame_data, text = my_prog.new_hire, relief=tk.SUNKEN)
    lbl_new_hire.grid(row=5, column=1, sticky="ew")

    # Insert button to select last new hire data file location
    btn_new_hire = tk.Button(frame_data, text="Select File", command=select_new_hire)
    btn_new_hire.grid(row=5, column=2, sticky="ew", padx=5, pady=5)

    ###################################
    #
    # CREATE USER DISPLAY AT BOTTOM
    #
    ###################################

    # Used to show draft report before exporting
    tabControl = ttk.Notebook(window)
    tab1 = ttk.Frame(tabControl)
    tab2 = ttk.Frame(tabControl)
    tab3 = ttk.Frame(tabControl)
    tab4 = ttk.Frame(tabControl)
    
    tabControl.add(tab1, text="Bench Data")
    tabControl.add(tab2, text="Labor Data")
    tabControl.add(tab3, text="Generated Report")
    tabControl.add(tab4, text="Employee to Job Matches")
    tabControl.grid(row=1, column=0, sticky="nsew", columnspan=3)

    # Create tab for bench report
    tab1_txt = tk.Text(tab1,width=200, xscrollcommand="v.set", yscrollcommand="v.set" ,wrap="none")
    tab1_txt.insert('1.0', "Import Data to display here.")
    tab1_txt.grid(row=1, column=0, sticky="nsew")
    tab1_txt.config(state='disabled')

    # Create tab for labor report
    tab2_txt = tk.Text(tab2,width=200, xscrollcommand="v.set", yscrollcommand="v.set" ,wrap="none")
    tab2_txt.insert('1.0', "Import Data to display here.")
    tab2_txt.grid(row=1, column=0, sticky="nsew")
    tab2_txt.config(state='disabled')

    # Create tab for Last SASM
    tab3_txt = tk.Text(tab3,width=200, xscrollcommand="v.set", yscrollcommand="v.set" ,wrap="none")
    tab3_txt.insert('1.0', "Press generate report to display here.")
    tab3_txt.grid(row=1, column=0, sticky="nsew")
    tab3_txt.config(state='disabled')

    # Create tab for Paired Report
    tab4_txt = tk.Text(tab4,width=200, xscrollcommand="v.set", yscrollcommand="v.set" ,wrap="none")
    tab4_txt.insert('1.0', "Press pair openings to display here.")
    tab4_txt.grid(row=1, column=0, sticky="nsew")
    tab4_txt.config(state='disabled')
    
    window.mainloop()
