import streamlit as st
import pandas as pd

import backend
import sasm_globals
import SASM_v4 as sa

from io import StringIO
from openpyxl import load_workbook

# Used to store location to data files
DATA_DIR = sasm_globals.DATA_DIR
DB_FILE = sasm_globals.DB_FILE
NEW_HIRE = sasm_globals.NEW_HIRE # Number of days considered a new Hire
VAR_TO_TARGET = sasm_globals.VAR_TO_TARGET # Allowed variance to target
SUSP_ABS = sasm_globals.SUSP_ABS
P_DAYS = sasm_globals.periods['1st'][1]

##########################################
#
# IMPORT BENCH DATA
#
###########################################
def import_bench(user, mbt):

    st.subheader("Step 1:")
    
    # Allow user to upload a .csv file
    cur_bench = st.file_uploader("Select the current Bench.csv file.",key="cur_bench", type=[".csv"], help="Upload Bench Data")

    # if user selected a file, update the file name
    if cur_bench != None:
        user.bench_path = cur_bench.name
        
        # Check to see if .csv file is in utf-8 format
        try:
            mbt.df_bench = pd.read_csv(cur_bench)
            
        # Check if .csv file is in utf-16 format
        except Exception as e1:
            try:
                stringio = StringIO(cur_bench.getvalue().decode("utf-16"))
                mbt.df_bench = pd.read_csv(stringio, sep='\t')
                
            # Tell user if file is in incorrect format
            except Exception as e2:
                st.warning(f"{cur_bench.name} is not in a valid .csv format")
                
        # Verify that data is in correct format
        import_col = mbt.df_bench.columns.tolist()
        chk_col = ["Empl Name & ID", "DL+Absc Hrs Variance to ITM Target %","Month of Hire Date"]
        
        # If not all correct columns, return error
        if not all(x in import_col for x in chk_col):
            st.warning("Bench data needs fields: 'Empl Name & ID' and 'DL+Absc Hrs Variance to ITM Target %'")
            user.bench_path = "No File Selected"
            return
            
        # Convert Month of hire to datetime
        mbt.df_bench['Month of Hire Date'] = pd.to_datetime(mbt.df_bench['Month of Hire Date'])
        mbt.df_bench['Month of Hire Date'] = mbt.df_bench['Month of Hire Date'].dt.strftime('%B %Y')
     

##########################################
#
# IMPORT LAST BENCH DATA
#
###########################################
def import_last_bench(user, mbt):

    st.subheader("Step 2:")
    last_bench = st.file_uploader("Select the last Bench.csv file.",key="last_bench", type=[".csv"], help="Upload Previous Bench Data")

    # if user selected a file, update the file name
    if last_bench != None:
        user.last_bench = last_bench.name
        
        # Check to see if .csv file is in utf-8 format
        try:
            mbt.df_last_bench = pd.read_csv(last_bench)
            
        # Check if .csv file is in utf-16 format
        except Exception as e1:
            try:
                stringio = StringIO(last_bench.getvalue().decode("utf-16"))
                mbt.df_last_bench = pd.read_csv(stringio, sep='\t')
                
            # Tell user if file is in incorrect format
            except Exception as e2:
                st.warning(f"{last_bench.name} is not in a valid .csv format")
                
        # Verify that data is in correct format
        import_col = mbt.df_last_bench.columns.tolist()
        chk_col = ["Empl Name & ID", "DL+Absc Hrs Variance to ITM Target %","Month of Hire Date"]
        
        # If not all correct columns, return error
        if not all(x in import_col for x in chk_col):
            st.warning("Bench data needs fields: 'Empl Name & ID' and 'DL+Absc Hrs Variance to ITM Target %'")
            user.last_bench = "No File Selected"
            return

##########################################
#
# IMPORT LABOR DATA
#
###########################################
def import_labor_data(user, mbt):
            
    st.subheader("Step 3:")
    labor = st.file_uploader("Select the Labor.csv file.",key="labor",
            type=['.csv'], help="Upload Labor Data")

    # if user selected a file, update the file name
    if labor != None:
        user.labor_path = labor.name
        
        # Check to see if .csv file is in utf-8 format
        try:
            mbt.df_labor = pd.read_csv(labor)
            
        # Check if .csv file is in utf-16 format
        except Exception as e1:
            try:
                stringio = StringIO(labor.getvalue().decode("utf-16"))
                mbt.df_labor = pd.read_csv(stringio, sep='\t')
                
            # Tell user if file is in incorrect format
            except Exception as e2:
                st.warning(f"{labor.name} is not in a valid .csv format")
                
        # Ensure data sources are in the correct format
        import_col = mbt.df_labor.columns.tolist()
        chk_col = ["Max. Hire Dt","Emplid","Billability Variance to Target","Billability","Suspense Amount","DL $ Target ","Total Absence Amount"]
        if not all(x in import_col for x in chk_col):
            st.warning("Labor data needs fields: 'Max. Hire Dt','Emplid','Billability Variance to Target','Billability','Suspense Amount','DL $ Target ','Total Absence Amount'")
            user.labor_path = "No File Selected"
            return

##########################################
#
# IMPORT LAST SASM DATA
#
###########################################
def import_last_sasm(user, mbt):

    st.subheader("Step 4:")
    last_sasm = st.file_uploader("Select the Last SASM.xlsx file.",key="last_sm",
            type=['.xlsx'], help="Upload Last SASM data")
            
    # if user selected a file, update the file name
    if last_sasm != None:
        user.last_sasm = last_sasm.name
        
        book = load_workbook(last_sasm)
        ws = book.worksheets[0]
        for cell in ws["C"]:
            if cell.value is not None:
                skip_row = cell.row - 1
                break
                
        # Check to see if .csv file is in utf-8 format
        try:
            mbt.df_last_sasm = pd.read_excel(last_sasm,skiprows=skip_row, engine="openpyxl")
            
        # Check if .csv file is in utf-16 format
        except Exception as e1:
            st.warning(f"{last_sasm.name} is not in a valid .xlsx format")
                
        # Ensure data sources are in the correct format
        import_col = mbt.df_last_sasm.columns.tolist()
        chk_col = ["SASM Notes", "SASM RM Status", "Anticipated Start Date", "Activity Comments"] 
        if not all(x in import_col for x in chk_col):
            st.warning("Last SASM data needs fields: 'SASM Notes', 'SASM RM Status', and 'Anticipated Start Date'")
            # user.last_sasm = "No File Selected"
            # return


##########################################
#
# DISPLAY SIDE BAR
#
###########################################
def side_bar(user,mbt):
    
    # Create a status table using a stoplight chart
    # format stoplights
    red = """<style>
            .dot1 {
              height: 25px;
              width: 25px;
              background-color: rgb(255,0,0);
              border-radius: 50%;
              display: inline-block;
            }
        </style><span class='dot1'></span>"""
        
    green = """<style>
            .dot2 {
              height: 25px;
              width: 25px;
              background-color: rgb(0,255,0);
              border-radius: 50%;
              display: inline-block;
            }
        </style><span class='dot2'></span>"""
        
    
    # Check to see status of files
    if user.bench_path == "No File Selected":
        b_col = red

    else:
        b_col = green
  
    if user.last_bench == "No File Selected":
        lb_col = red
    else:
        lb_col = green
       
    if user.labor_path == "No File Selected":
        la_col = red
    else:
        la_col = green
     
    if user.last_sasm == "No File Selected":
        ub_col = red
    else:
        ub_col = green
        
    
    st.markdown(
    f"""
    <table>
        <tr>
            <th>Data</th>
            <th>Status</th>
        </tr>
        <tr>
            <th>Bench Data:</th>
            <th>{b_col}</th>
        </tr>
        <tr>
            <th>Last Bench:</th>
            <th>{lb_col}</th>
        </tr>
        <tr>
            <th>Labor Data:</th>
            <th>{la_col}</th>
        </tr>
        <tr>
            <th>Last SASM:</th>
            <th>{ub_col}</th>
        </tr>
    </table
    <br>
    <br>
    """, unsafe_allow_html=True)
    
    # Display to user if ready to generate bench report
    if la_col == lb_col == b_col == ub_col == green:
        st.markdown("""
        <p style="font-family:Courier; color:Green; font-size: 20px;">Ready To Generate</p>
        """, unsafe_allow_html=True)
        st.button(label="Generate Report", on_click=lambda:generate_report(user,mbt))
    
        # Export report if ready
        if not mbt.df_final_report.empty:
            with open(user.file_path,'rb') as f:
                st.download_button("Download Bench Report", f, file_name='bench.xlsx')
            #st.button(label="Export Report", on_click=lambda:sa.export_btn(mbt,user))
    else:
        st.markdown("""
        <p style="font-family:Courier; color:Red; font-size: 20px;">Select Files</p>
        """, unsafe_allow_html=True)


    
def generate_report(user, mbt):
    """
    generate_report function does ALL the data analysis.  It uses the bench and labor
    dataframes and runs the following deciion tree:
        1. If member is on bench list verify:
            a. No approved leave
            b. No suspended charges
            c. Not in a non-billable billet
            d. Not a hew hire (< 6 months)
        2. If member is not on bench list, check associated labor data and verify:
            a. Member billed required hours, s.t. constraints in paragraph 1.

        The results are stored in the final_bench_report dataframe contained in
        the my_billable_tracker object

    :param my_billable_tracker: my_billable_tracker is an object that stores
                                two pandas dataframes:
                                    1. bench data
                                    2. labor data
                                    3. final_bench_report
 
    :return: None
    """

    # Ensure the data was inputed prior to generating the report
    # import_data(my_billable_tracker)

    # Pre-process the data, e.g., index, types, etc
    backend.pre_process(mbt)
    
    #########################################
    #
    # RUN DECISION TREE
    #
    #########################################

    # 1st decition point
    # Is bench data < 4 business days of this reporting period
    use_bench_only = backend.dt_less_4_bd(mbt,user)
    # use_bench_only = False
    
    # If we pass dt point 1, continue with the rest of the tree
    if use_bench_only:
        st.warning(f"Within {P_DAYS} days of reporting period, only using data from {user.bench_path}.")

    else:

        # This decision point check to see if employee is billabile
        backend.dt_isBillable(mbt)

        # This decision points checks to see(SUSP + ABS > 0.75) and (NOT on prior period bench)
        backend.dt_pt_2(mbt)

    # (DL + ABS HRS VAR to ITM target % < -0.50) && (Billability Target > 0.50)?
    backend.dt_pt_3(mbt)
    
    # If employee was on official SASM, do NOT move to watchlist
    backend.dt_pt_4(mbt)


    #########################################
    #
    # UPDATE GUI
    #
    #########################################
    # Remove Senior Associates and Principals
    mbt.df_consolidated.loc[(mbt.df_consolidated['Level '] == "Senior Associate"), "Notes"] = "Senior Associate - Removed from tracker"
    mbt.df_consolidated.loc[(mbt.df_consolidated['Level '] == "Principal/Director"), "Notes"] = "Princial - Removed from tracker"
    mbt.df_consolidated.loc[(mbt.df_consolidated['Level '] == "Senior Associate"), "on_bench"] = False
    mbt.df_consolidated.loc[(mbt.df_consolidated['Level '] == "Principal/Director"), "on_bench"] = False

    # Generate the final bench report and watchlist
    mbt.df_final_report = mbt.df_consolidated.loc[mbt.df_consolidated['on_bench'] == True]
    mbt.df_watch_list = mbt.df_consolidated.loc[mbt.df_consolidated['on_bench'] == "watch"]

    # Scrub 'Senior Associate' and "Principal/Director"
    mbt.df_consolidated = mbt.df_consolidated.loc[mbt.df_consolidated['Level '] != "Senior Associate"]
    mbt.df_consolidated = mbt.df_consolidated.loc[mbt.df_consolidated['Level '] != "Principal/Director"]

    # Catagorize the new hires
    # Convert Month of hire to datetime
    
    # df['Date']= pd.to_datetime(df['Date'])
    mbt.df_consolidated['Max. Hire Dt'] = pd.to_datetime(mbt.df_consolidated['Max. Hire Dt'])

    mbt.df_new_hire = mbt.df_consolidated.loc[mbt.df_consolidated['Max. Hire Dt'] + pd.DateOffset(days=NEW_HIRE) >= pd.Timestamp.today()]

    # Update column names to match desiered output.
    # Rename columns to match previous report
    mbt.df_new_hire = mbt.df_new_hire.rename(columns={"Full Legal Name":"Name", "Job Leader":"CM as of Hire Date"})

    cols = ["Name", "Hire Date", "Job Requisition","CM as of Hire Date", "Principal", "SASM POC",
            "Management Level", "Location", "Opportunities (Contract field)", "Requisition Type", "Job Family",
             "Clearance as confirmed by CM"
    ]
    mbt.df_new_hire = mbt.df_new_hire.reindex(columns=cols)
    
    sa.export_btn(mbt,user)

#    # Check to see if there is new hire data in the previous SASM.
#    # If there is, import the old data
#    try:
#        book = load_workbook(lbl_last_sasm["text"])
#        ws = book["New Hires"]
#
#        # Check to see if there are blank rows to skip
#        for cell in ws["C"]:
#            if cell.value is not None:
#                skip_row = cell.row - 1
#                break
#
#        # Load ws into temp df
#        tmp_df = pd.read_excel(lbl_last_sasm["text"], sheet_name="New Hires",
#                                skiprows=skip_row, engine="openpyxl")
#        tmp_df = pd.DataFrame(data=tmp_df)
#        tmp_df.set_index("Employee ID", inplace=True)
#
#        # Copy applicable data from last_sasm to new hires df
#        my_billable_tracker.df_new_hire.loc[my_billable_tracker.df_new_hire.index.isin(tmp_df.index),
#                                            "Job Requisition"] = tmp_df['Job Requisition']
#        my_billable_tracker.df_new_hire.loc[my_billable_tracker.df_new_hire.index.isin(tmp_df.index),
#                                            "Opportunities (Contract field)"] = tmp_df["Opportunities (Contract field)"]
#        my_billable_tracker.df_new_hire.loc[my_billable_tracker.df_new_hire.index.isin(tmp_df.index),
#                                            "Requisition Type"] = tmp_df["Requisition Type"]
#        my_billable_tracker.df_new_hire.loc[my_billable_tracker.df_new_hire.index.isin(tmp_df.index),
#                                            "Clearance as confirmed by CM"] = tmp_df["Clearance as confirmed by CM"]
#
#        # my_billable_tracker.df_new_hire.to_csv("dump.csv")
#    except Exception as e:
#        # print(e)
#        tk.messagebox.showwarning("No 'New Hires' tab in previous SASM reprt", "No 'New Hires' tab in previous SASM reprt")
#
